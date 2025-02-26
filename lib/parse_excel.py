from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import sys
def is_a_workout_day(row):
  return (row[0].value is not None and row[1].value is not None) and 'Set 1' not in [x.value for x in row]

workbook = load_workbook(filename=sys.argv[1])
#load_workbook(filename='/Users/chabgood/projects/calimove_tracker/spec/fixtures/files/test_upload.xlsx')
sheet = workbook.active
primary_row = list([row for row in sheet.iter_rows(values_only=True) if 'Phase' in row][0])
if primary_row[-1] is None:
  del primary_row[-1]
num_of_cols = len(primary_row)
num_of_sets = len(primary_row[9:])
cali_move = []
for row in sheet.rows:
  row = list(row)
  if len(row) != num_of_cols:
    del row[-1]
  if is_a_workout_day(row):
    orig_row = row
    row = row[9:]
    sets = 0
    for column in row:
      if column.fill.fgColor.rgb == '00000000' or column.fill.fgColor.rgb == 'FFD9D9D9':
        sets += 1
    cali_move.append(sets)

print(cali_move)
